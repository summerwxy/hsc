#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from colorama import init

import urllib.request
from lxml import etree, cssselect
import xlsxwriter


def getDecks(url):
  print('\033[1;31m>> start get decks list\033[m')
  decks = []
  string = urllib.request.urlopen(url).read()
  parser = etree.HTMLParser()
  html = etree.fromstring(string, parser)
  select = cssselect.CSSSelector(r'#decks td')
  items = select(html)
  i = 0
  while i < len(items): 
    deck = {}
    # Deck Name
    a = items[i].find(r'div/span/a')
    deck['name'] = a.text.strip()
    deck['url'] = 'http://www.hearthpwn.com' + a.get('href')
    arena = items[i].get('class').find('t-arena-cell') != -1 and True or False
    print(arena)
    i = i + 1
    # Deck Type
    deck['type'] = items[i].text.strip()
    i = i + 1
    # Mana
    i = i + 1
    # Class
    deck['class'] = items[i].text.strip()
    i = i + 1
    # Rating 
    deck['rating'] = items[i].find(r'div').text.strip()
    i = i + 1
    # Views
    deck['views'] = items[i].text.strip()
    i = i + 1
    # Comments
    i = i + 1
    deck['comments'] = items[i].text.strip()
    # Cost
    deck['cost'] = items[i].text.strip()
    i = i + 1
    # Updated
    deck['updated'] = items[i].find(r'abbr').get('title')
    deck['patch'] = items[i].find(r'span').text.strip()
    i = i + 1
    # if arena deck pass
    if not arena:
      decks.append(deck)
      print(deck)
  return decks


def getDeck(decks):
  for deck in decks:
    print('\033[1;31m>> get deck: %s\033[m' % (deck['name']))
    string = urllib.request.urlopen(deck['url']).read()
    parser = etree.HTMLParser()
    html = etree.fromstring(string, parser)
    select = cssselect.CSSSelector(r'.listing-cards-tabular td')
    items = select(html)
    i = 0
    cards = []
    while i < len(items):
      card = {}
      # Name
      card['name'] = items[i].find(r'b/a').text.strip()
      card['count'] = items[i].find(r'b').tail.strip()[-1]
      i = i + 1
      # Cost
      card['cost'] = items[i].text.strip()
      i = i + 1
      cards.append(card)
    deck['cards'] = cards
  return decks

def getMyCards():
  print('\033[1;31m>> read excel data\033[m')
  cards = {}
  wb = load_workbook(filename='cards-result.xlsx')
  ws = wb['Sheet1']
  for row in ws.rows:
    name = row[29].value
    have = row[3].value
    cname = row[10].value
    if have != None:
      cards[name] = {'have': have, 'cname': cname}
  return cards

def write2excel(decks, cards):
  print('\033[1;31m>> write to excel\033[m')
  workbook = xlsxwriter.Workbook('hearthpwn-result.xlsx')
  green = workbook.add_format({'bg_color': '#c2d69a', 'border': 1})
  red = workbook.add_format({'bg_color': '#e6b9b8', 'border': 1})
  info_title = [['Deck Name', 'name'], ['URL', 'url'], ['Type', 'type'], ['Class', 'class'], ['Rating', 'rating'], ['Views', 'views'], ['Comments', 'comments'], ['Cost', 'cost'], ['Updated', 'updated'], ['Patch', 'patch']]

  for deck in decks:
    worksheet = workbook.add_worksheet()
    for i in range(len(info_title)):
      worksheet.write(i, 0, info_title[i][0])
      worksheet.write(i, 1, deck[info_title[i][1]])

    i = len(info_title) + 3
    worksheet.write(i, 0, 'Name')
    worksheet.write(i, 1, 'Name')
    worksheet.write(i, 2, 'Cost')
    worksheet.write(i, 3, 'Count')
    worksheet.write(i, 4, 'Have')
    i = i + 1
    for card in deck['cards']:
      worksheet.write(i, 0, card['name'])
      worksheet.write(i, 1, cards[card['name']]['cname'])
      worksheet.write(i, 2, int(card['cost']))
      count = int(card['count'])
      worksheet.write(i, 3, count)
      have = int(cards[card['name']]['have'])
      worksheet.write(i, 4, have)
      worksheet.set_row(i, 13.5, (have >= count) and green or red)
      i = i + 1
    ws = [25, 20, 10, 10, 10]
    for i in range(len(ws)):
      worksheet.set_column(i, i, ws[i])


if __name__ == '__main__':
  init() # use color console
  url = input('please enter url: ')
  url = url or 'http://www.hearthpwn.com/decks'
  decks = getDecks(url)
  decks = getDeck(decks)
  cards = getMyCards()
  write2excel(decks, cards)


